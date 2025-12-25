#提交任务
from typing import List, Optional
from pydantic import BaseModel
from DrissionPage import Chromium
import openpyxl,asyncio,datetime,re,pickle
from pathlib import Path
import win32com.client as win32

BROWSER=Chromium()
URL='https://www.jabdc.com'
CashBusinessInfoName='businessInfo.pkl'

class BusinessInfo(BaseModel):
    business_type:str
    customers:List[dict]

    loan_type:Optional[str]=None
    loan_manger:Optional[str]=None
    loan_amount:Optional[int]=None
    loan_contract:Optional[str]=None
    loan_tream:Optional[dict]=None
    loan_guarantee:Optional[str]=None

    certificate:Optional[List[str]]=None
    #义务人开发商信息
    obligor:Optional[dict]=None
    #购房备案合同号
    purchase_contract:Optional[str]=None
    #图片路径
    img_path:Optional[str]=None



#将数据写入excel新的一行(openpyxl)
async def writeBusinessInfoToExcel(businessInfo:BusinessInfo,file_path:str):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.worksheets[1]
    row=sheet.max_row+1
    sheet.cell(row=row, column=10, value=businessInfo.business_type)
    sheet.cell(row=row, column=2, value=businessInfo.customers[0]["name"]).alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet.cell(row=row, column=3, value=businessInfo.customers[0]["id"])
    sheet.cell(row=row, column=4, value=businessInfo.customers[0]["phone"])
    if len(businessInfo.customers)>1:
        customers_list=businessInfo.customers[1:]
        name_list=list(map(lambda x:x["name"],customers_list))
        id_list=list(map(lambda x:x["id"],customers_list))
        phone_list=list(map(lambda x:x["phone"],customers_list))
        sheet.cell(row=row, column=5, value='\n'.join(name_list)).alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet.cell(row=row, column=6, value='\n'.join(id_list))
        sheet.cell(row=row, column=7, value='\n'.join(phone_list))
    if businessInfo.certificate:
        sheet.cell(row=row, column=8, value='\n'.join(businessInfo.certificate))
    if businessInfo.loan_contract:
        sheet.cell(row=row, column=9, value=businessInfo.loan_contract)
    sheet.cell(row=row, column=11, value=datetime.datetime.today().strftime('%Y/%m/%d'))
    if businessInfo.loan_type:
        sheet.cell(row=row, column=13, value=businessInfo.loan_type)
    if businessInfo.loan_manger:
        sheet.cell(row=row, column=14, value=businessInfo.loan_manger)

    workbook.save(file_path)
    pass


#将数据实时写入wps表格（pywin32）
async def writeBusinessInfoToWps(businessInfo:BusinessInfo,file_path:str):
    try:    
        wps = win32.GetObject(Class='ket.Application')
        print('已成功连接wps表格')
    except:
        wps = win32.Dispatch('ket.Application')
        print('未找到已打开的WPS表格，将启动新实例。')
    wps.Visible = True
    wb=''
    for workbook in wps.Workbooks:
        if Path(workbook.FullName).name == Path(file_path).name:
            workbook.Activate()
            wb=workbook
            break
    if wb=='':
        try:
            wb=wps.Workbooks.Open(file_path)
        except:
            print(f'未找到指定文件:{file_path}')
            return
    if '合并登记' in businessInfo.business_type:
        sheet=wb.Worksheets(1)
        sheet.Activate()
        new_row=sheet.UsedRange.Rows.Count+1
        sheet.Cells(new_row, 4).Value=businessInfo.customers[0]["name"]
        sheet.Cells(new_row, 5).Value=businessInfo.customers[0]["id"]
        sheet.Cells(new_row, 6).Value=businessInfo.customers[0]["phone"]
        if len(businessInfo.customers)>1:
            customers_list=businessInfo.customers[1:]
            name_list=list(map(lambda x:x["name"],customers_list))
            id_list=list(map(lambda x:x["id"],customers_list))
            phone_list=list(map(lambda x:x["phone"],customers_list))
            sheet.Cells(new_row, 7).Value='\n'.join(name_list)
            sheet.Cells(new_row, 8).Value='\n'.join(id_list)
            sheet.Cells(new_row, 9).Value='\n'.join(phone_list)
        if businessInfo.purchase_contract:
            sheet.Cells(new_row, 2).Value=businessInfo.purchase_contract
        if businessInfo.loan_contract:
            sheet.Cells(new_row, 13).Value=businessInfo.loan_contract
        sheet.Cells(new_row, 14).Value=datetime.datetime.today().strftime('%Y/%m/%d')
        if businessInfo.obligor:
            sheet.Cells(new_row, 10).Value=businessInfo.obligor["name"]
            sheet.Cells(new_row, 11).Value=businessInfo.obligor["id"]
            sheet.Cells(new_row, 12).Value=businessInfo.obligor["phone"]
        if businessInfo.loan_manger:
            sheet.Cells(new_row, 16).Value=businessInfo.loan_manger
    else:
        sheet=wb.Worksheets(2)
        new_row=sheet.UsedRange.Rows.Count+1
        sheet.Activate()
        sheet.Cells(new_row, 10).Value=businessInfo.business_type
        sheet.Cells(new_row, 2).Value=businessInfo.customers[0]["name"]
        sheet.Cells(new_row, 3).Value=businessInfo.customers[0]["id"]
        sheet.Cells(new_row, 4).Value=businessInfo.customers[0]["phone"]
        if len(businessInfo.customers)>1:
            customers_list=businessInfo.customers[1:]
            name_list=list(map(lambda x:x["name"],customers_list))
            id_list=list(map(lambda x:x["id"],customers_list))
            phone_list=list(map(lambda x:x["phone"],customers_list))
            sheet.Cells(new_row, 5).Value='\n'.join(name_list)
            sheet.Cells(new_row, 6).Value='\n'.join(id_list)
            sheet.Cells(new_row, 7).Value='\n'.join(phone_list)
        if businessInfo.certificate:
            sheet.Cells(new_row, 8).Value='\n'.join(businessInfo.certificate)
        if businessInfo.loan_contract:
            sheet.Cells(new_row, 9).Value=businessInfo.loan_contract
        sheet.Cells(new_row, 11).Value=datetime.datetime.today().strftime('%Y/%m/%d')
        if businessInfo.loan_type:
            sheet.Cells(new_row, 13).Value=businessInfo.loan_type
        if businessInfo.loan_manger:
            sheet.Cells(new_row, 14).Value=businessInfo.loan_manger
    
    wb.Save()
    


#缓存businessinfo到图片文件夹
async def cacheBusinessInfo(businessInfo:BusinessInfo):
    cache_path=Path(businessInfo.img_path)/CashBusinessInfoName
    with open(cache_path,'wb') as f:
        pickle.dump(businessInfo,f)
    
#读取businessinfo到对象
async def readBusinessInfoFromCache(path:str):
    cache_path=Path(path)/CashBusinessInfoName
    if cache_path.exists():
        with open(cache_path,'rb') as f:
            businessInfo=pickle.load(f)
            return businessInfo
    else:
        return None
    
#将数据写入sqlite


#网页操作 - 选择业务类型并进入
async def webSelectType(business_type:str,certificate:Optional[List[str]]=None):
    new_certificate=[]
    if certificate:
        for cert in certificate:
           res=re.match(r".+?[(（](\d+)[)）].+?第(\d+)号",cert)
           if res:
            new_certificate.append([res.group(1),res.group(2)])
    #print(new_certificate)
    tab=BROWSER.get_tab(url=URL)
    BROWSER.activate_tab(tab)
    tab.ele('tag=div@text():待提交').click()
    tab.ele('tag=div@text():新增').click()
    tab.wait.ele_displayed('@placeholder:请选择申请大类',timeout=5)
    '''
    match business_type:
        case "一般抵押":
            tab.ele('@placeholder:请选择申请大类',timeout=3).click()
            tab.ele('text:房屋抵押登记',timeout=3).click()
            tab.ele('@placeholder:请选择申请小类',timeout=3).click()
            tab.ele('text:一般抵押',timeout=3).click()
            tab.ele('@placeholder:请选择证书类型',timeout=3).click()
            if len(new_certificate)>0:
                tab.ele('text:不动产权证号',timeout=3).click()
            else:
                tab.ele('text:房产证',timeout=3).click()
            
        case "最高额抵押":
            tab.ele('@placeholder:请选择申请大类',timeout=3).click()
            tab.ele('text:房屋抵押登记',timeout=3).click()
            tab.ele('@placeholder:请选择申请小类',timeout=3).click()
            tab.ele('text:最高额抵押',timeout=3).click()
            tab.ele('@placeholder:请选择证书类型',timeout=3).click()
            if len(new_certificate)>0:
                tab.ele('text:不动产权证号',timeout=3).click()
            else:
                tab.ele('text:房产证',timeout=3).click()

        case "合并登记（预告抵押）":
            tab.ele('@placeholder:请选择申请大类',timeout=3).click()
            tab.ele('text:合并登记',timeout=3).click()
            tab.ele('@placeholder:请选择申请小类',timeout=3).click()
            tab.ele('text:转移抵押预告合一登记',timeout=3).click()
            pass
        case "预告抵押（仅预告抵押）":
            tab.ele('@placeholder:请选择申请大类',timeout=3).click()
            tab.ele('text:预告登记',timeout=3).click()
            tab.ele('@placeholder:请选择申请小类',timeout=3).click()
            tab.ele('text:预售商品房抵押权预告登记',timeout=3).click()
            tab.ele('text:预告买卖证明号',timeout=3).after('tag=input',index=1).input(new_certificate[0][0],clear=True)
            tab.ele('text:预告买卖证明号',timeout=3).after('tag=input',index=2).input(new_certificate[0][1],clear=True)
            pass
        case "转本登记（预告转正式）":
            tab.ele('@placeholder:请选择申请大类',timeout=3).click()
            tab.ele('text:转本登记',timeout=3).click()
            tab.ele('@placeholder:请选择申请小类',timeout=3).click()
            tab.ele('text:预售商品房抵押权预告转本登记',timeout=3).click()
            tab.ele('@placeholder:请选择证书类型',timeout=3).click()
            if len(new_certificate)>0:
                tab.ele('text:不动产登记证明',timeout=3).click()
                tab.ele('tag=label@text():不动产证明',timeout=3).after('tag=input',index=1).input(new_certificate[0][0],clear=True)
                tab.ele('tag=label@text():不动产证明',timeout=3).after('tag=input',index=2).input(new_certificate[0][1],clear=True)
            else:
                tab.ele('text:他项权证',timeout=3).click()
                tab.ele('@placeholder:请输入他项权证',timeout=3).input(certificate[0],clear=True)
            pass
        case "预告解押":
            tab.ele('@placeholder:请选择申请大类',timeout=3).click()
            tab.ele('text:预告登记',timeout=3).click()
            tab.ele('@placeholder:请选择申请小类',timeout=3).click()
            tab.ele('text:抵押预告注销登记',timeout=3).click()
            tab.ele('@placeholder:请选择证书类型',timeout=3).click()
            if len(new_certificate)>0:
                tab.ele('text:预告抵押证明',timeout=3).click()
                tab.ele('tag=label@text():预告抵押证明号',timeout=3).after('tag=input',index=1).input(new_certificate[0][0],clear=True)
                tab.ele('tag=label@text():预告抵押证明号',timeout=3).after('tag=input',index=2).input(new_certificate[0][1],clear=True)
            else:
                tab.ele('text:他项权证',timeout=3).click()
                tab.ele('@placeholder:请输入他项权证',timeout=3).input(certificate[0],clear=True)
            pass
        case "解押":
            tab.ele('@placeholder:请选择申请大类',timeout=3).click()
            tab.ele('text:房屋抵押注销登记',timeout=3).click()
            tab.ele('@placeholder:请选择申请小类',timeout=3).click()
            tab.ele('text:房屋抵押注销登记',timeout=3).click()
            tab.ele('@placeholder:请选择证书类型',timeout=3).click()
            if len(new_certificate)>0:
                tab.ele('text:不动产登记证明',timeout=3).click()
                tab.ele('tag=label@text():不动产证明',timeout=3).after('tag=input',index=1).input(new_certificate[0][0],clear=True)
                tab.ele('tag=label@text():不动产证明',timeout=3).after('tag=input',index=2).input(new_certificate[0][1],clear=True)
            else:
                tab.ele('text:他项权证',timeout=3).click()
                tab.ele('@placeholder:请输入他项权证',timeout=3).input(certificate[0],clear=True)
            pass
    '''
    #兼容python3.8写法
    if business_type=="一般抵押":
        tab.ele('@placeholder:请选择申请大类',timeout=3).click()
        tab.ele('text:房屋抵押登记',timeout=3).click()
        tab.ele('@placeholder:请选择申请小类',timeout=3).click()
        tab.ele('text:一般抵押',timeout=3).click()
        tab.ele('@placeholder:请选择证书类型',timeout=3).click()
        if len(new_certificate)>0:
            tab.ele('text:不动产权证号',timeout=3).click()
        else:
            tab.ele('text:房产证',timeout=3).click()
            
    elif business_type=="最高额抵押":
        tab.ele('@placeholder:请选择申请大类',timeout=3).click()
        tab.ele('text:房屋抵押登记',timeout=3).click()
        tab.ele('@placeholder:请选择申请小类',timeout=3).click()
        tab.ele('text:最高额抵押',timeout=3).click()
        tab.ele('@placeholder:请选择证书类型',timeout=3).click()
        if len(new_certificate)>0:
            tab.ele('text:不动产权证号',timeout=3).click()
        else:
            tab.ele('text:房产证',timeout=3).click()

    elif business_type=="合并登记（预告抵押）":
        tab.ele('@placeholder:请选择申请大类',timeout=3).click()
        tab.ele('text:合并登记',timeout=3).click()
        tab.ele('@placeholder:请选择申请小类',timeout=3).click()
        tab.ele('text:转移抵押预告合一登记',timeout=3).click()
        pass
    elif business_type=="预告抵押（仅预告抵押）":
        tab.ele('@placeholder:请选择申请大类',timeout=3).click()
        tab.ele('text:预告登记',timeout=3).click()
        tab.ele('@placeholder:请选择申请小类',timeout=3).click()
        tab.ele('text:预售商品房抵押权预告登记',timeout=3).click()
        tab.ele('text:预告买卖证明号',timeout=3).after('tag=input',index=1).input(new_certificate[0][0],clear=True)
        tab.ele('text:预告买卖证明号',timeout=3).after('tag=input',index=2).input(new_certificate[0][1],clear=True)
        pass
    elif business_type=="转本登记（预告转正式）":
        tab.ele('@placeholder:请选择申请大类',timeout=3).click()
        tab.ele('text:转本登记',timeout=3).click()
        tab.ele('@placeholder:请选择申请小类',timeout=3).click()
        tab.ele('text:预售商品房抵押权预告转本登记',timeout=3).click()
        tab.ele('@placeholder:请选择证书类型',timeout=3).click()
        if len(new_certificate)>0:
            tab.ele('text:不动产登记证明',timeout=3).click()
            tab.ele('tag=label@text():不动产证明',timeout=3).after('tag=input',index=1).input(new_certificate[0][0],clear=True)
            tab.ele('tag=label@text():不动产证明',timeout=3).after('tag=input',index=2).input(new_certificate[0][1],clear=True)
        else:
            tab.ele('text:他项权证',timeout=3).click()
            tab.ele('@placeholder:请输入他项权证',timeout=3).input(certificate[0],clear=True)
        pass
    elif business_type=="预告解押":
        tab.ele('@placeholder:请选择申请大类',timeout=3).click()
        tab.ele('text:预告登记',timeout=3).click()
        tab.ele('@placeholder:请选择申请小类',timeout=3).click()
        tab.ele('text:抵押预告注销登记',timeout=3).click()
        tab.ele('@placeholder:请选择证书类型',timeout=3).click()
        if len(new_certificate)>0:
            tab.ele('text:预告抵押证明',timeout=3).click()
            tab.ele('tag=label@text():预告抵押证明号',timeout=3).after('tag=input',index=1).input(new_certificate[0][0],clear=True)
            tab.ele('tag=label@text():预告抵押证明号',timeout=3).after('tag=input',index=2).input(new_certificate[0][1],clear=True)
        else:
            tab.ele('text:他项权证',timeout=3).click()
            tab.ele('@placeholder:请输入他项权证',timeout=3).input(certificate[0],clear=True)
        pass
    elif business_type=="解押":
        tab.ele('@placeholder:请选择申请大类',timeout=3).click()
        tab.ele('tag=li@@text():房屋抵押注销登记',timeout=3).click()
        tab.ele('@placeholder:请选择申请小类',timeout=3).click()
        tab.eles('tag=li@@text():房屋抵押注销登记',timeout=3)[-1].click()
        tab.ele('@placeholder:请选择证书类型',timeout=3).click()
        if len(new_certificate)>0:
            tab.ele('text:不动产登记证明',timeout=3).click()
            tab.ele('tag=label@text():不动产证明',timeout=3).after('tag=input',index=1).input(new_certificate[0][0],clear=True)
            tab.ele('tag=label@text():不动产证明',timeout=3).after('tag=input',index=2).input(new_certificate[0][1],clear=True)
        else:
            tab.ele('text:他项权证',timeout=3).click()
            tab.ele('@placeholder:请输入他项权证',timeout=3).input(certificate[0],clear=True)
        pass

    tab.ele('text:确定',timeout=3).click()
    pass

#网页操作 - 录入信息
async def webInputInfo(businessInfo:BusinessInfo):
    if "解押" in businessInfo.business_type:
        return
    tab=BROWSER.get_tab(url=URL)
    #添加抵押人
    if businessInfo.business_type=="合并登记（预告抵押）":
        tab.ele('tag=h3@@text():业务类型',timeout=3).after('tag=input@placeholder:请输入合同编号',timeout=3).input(businessInfo.purchase_contract,clear=True)

        ele_add_customer=tab.ele('tag=h3@@text():权利人信息',timeout=3).after('tag=div@text():添加权利人')
        #添加开发商信息 义务人
        ele_add_obligor=tab.ele('tag=h3@@text():义务人信息',timeout=3).after('tag=div@text():添加义务人')
        if ele_add_obligor:
            ele_add_obligor.click()
            tab.ele('@placeholder:请选择证件种类',timeout=3).click()
            tab.eles('tag=li@@text()=统一社会信用代码@@class:el-select-dropdown__item',timeout=3)[-1].click()
            tab.ele('@placeholder:请输入姓名',timeout=3).input(businessInfo.obligor['name'],clear=True)
            tab.ele('@placeholder:请输入证件号',timeout=3).input(businessInfo.obligor['id'],clear=True)
            tab.ele('@placeholder:请输入联系方式',timeout=3).input(businessInfo.obligor['phone'],clear=True)
            ele_button=tab.ele('tag=span@text()=确认',timeout=3)
            ele_button.click()
    else:
        ele_add_customer=tab.ele('tag=h3@@text():抵押人信息',timeout=3).after('tag=div@@text():添加抵押人',timeout=3)

    if ele_add_customer:
        for customer in businessInfo.customers:
            ele_add_customer.click()
            tab.ele('@placeholder:请选择证件种类',timeout=3).click()
            tab.eles('tag=li@@text()=身份证@@class:el-select-dropdown__item',timeout=3)[-1].click()
            tab.ele('@placeholder:请输入姓名',timeout=3).input(customer['name'],clear=True)
            tab.ele('@placeholder:请输入身份证号',timeout=3).input(customer['id'],clear=True)
            tab.ele('@placeholder:请输入联系方式',timeout=3).input(customer['phone'],clear=True)
            ele_button=tab.ele('tag=span@text()=确认',timeout=3)
            ele_button.click()
            
    

    #添加抵押信息
    ele_loan_info=tab.ele('tag=div@@class:yw_con@@text():担保范围')
    #金额
    ele_loan_amount= ele_loan_info.ele('tag=label@|text():银行债权数额@|text():被担保债权数额').parent().ele('tag=input@placeholder:请输入债权数额',timeout=3)
    if ele_loan_amount:
        ele_loan_amount.input(businessInfo.loan_amount,clear=True)
    #抵押期限
    ele_loan_term_start= ele_loan_info.ele('tag=label@|text():银行债务履行期限@|text()=债务履行期限').parent().ele('tag=input@placeholder:开始日期',timeout=3)
    if ele_loan_term_start:
        ele_loan_term_start.input(businessInfo.loan_tream['start'],clear=True)
        ele_loan_term_start.after('tag=input@placeholder:结束日期',timeout=3).input(businessInfo.loan_tream['end'],clear=True)
        ele_loan_info.ele('tag=label@text():债务履行期限').click()
    #合同编号
    ele_loan_contract= ele_loan_info.ele('tag=div@|text():银行合同编号@|text()=合同编号').ele('tag=input@placeholder:请输入合同编号',timeout=3)
    if ele_loan_contract:
        ele_loan_contract.input(businessInfo.loan_contract,clear=True)
    
    #抵押顺位
    ele_loan_priority= ele_loan_info.ele('tag=input@placeholder:请选择抵押顺位',timeout=3)
    if ele_loan_priority:
        ele_loan_priority.click()
        tab.ele('tag=li@@text()=1@@class:el-select-dropdown__item',timeout=3).click()

    #担保范围
    ele_loan_guarantee= ele_loan_info.ele('tag=input@placeholder:请输入担保范围',timeout=3)
    if ele_loan_guarantee:
        ele_loan_guarantee.input(businessInfo.loan_guarantee,clear=True)

    # 仅预告抵押 抵押方式选一般抵押
    if businessInfo.business_type=="预告抵押（仅预告抵押）":
        ele_loan_type= ele_loan_info.ele('tag=input@placeholder:请选择抵押方式',timeout=3)
        if ele_loan_type:
            ele_loan_type.click()
            tab.ele('tag=li@@text()=一般抵押@@class:el-select-dropdown__item',timeout=3).click()
    # 转本登记（预告转正式）总金额贷款金额/0.7
    if businessInfo.business_type=="转本登记（预告转正式）":
        ele_loan_total= ele_loan_info.ele('tag=input@placeholder:请输入总金额',timeout=3)
        if ele_loan_total:
            ele_loan_total.input( round(float(businessInfo.loan_amount) /0.7,2),clear=True)


    #添加产权证号
    new_certificate=[]
    if businessInfo.certificate:
        for cert in businessInfo.certificate:
           res=re.match(r".+?[(（](\d+)[)）].+?第(\d+)号",cert)
           if res:
            new_certificate.append([res.group(1),res.group(2)])
    
    ele_zhTwo=tab.ele('tag=div@class=zhTwo',timeout=3)
    if ele_zhTwo:
        eles_certificate=ele_zhTwo.eles('tag=label@|text():产权证号@|text():房产证',timeout=3)
        for i in range(len(eles_certificate)):
            if len(new_certificate)>0:
                if len(new_certificate)>=i+1:
                    eles_certificate[i].after('tag=input',index=1).input(new_certificate[i][0],clear=True)
                    eles_certificate[i].after('tag=input',index=2).input(new_certificate[i][1],clear=True)
                else:
                    eles_certificate[i].after('tag=input',index=1).input(new_certificate[0][0],clear=True)
                    eles_certificate[i].after('tag=input',index=2).input(new_certificate[0][1],clear=True)
            else:
                if len(businessInfo.certificate)>=i+1:
                    eles_certificate[i].after('tag=input',index=1).input(businessInfo.certificate[i],clear=True)
                else:
                    eles_certificate[i].after('tag=input',index=1).input(businessInfo.certificate[0],clear=True)

#网页操作 - 上传影像
async def webUploadImg(businessInfo:BusinessInfo):
    tab=BROWSER.get_tab(url=URL)
    ele_annex_box=tab.ele('tag=div@class:annex_box',timeout=3)
    img_path=Path(businessInfo.img_path)
    for img_dir in img_path.iterdir():
        if img_dir.is_dir():
            img_list=list(map(lambda x: str(x),list(img_dir.iterdir())))
            '''
            不动产登记申请书
            不动产抵押登记询问笔录
            申请人身份证明
            委托函
            商品房预售合同
            不动产登记证明
            房产证
            借款合同
            抵押合同
            不动产抵押权注销证明
            关于抵押预告登记的约定
            结婚证或具结保证书
            其他
            '''
            if img_dir.name=="不动产登记申请书":
                ele_annex_box.ele('tag=li@@text():不动产登记申请书',timeout=3).ele('tag=div@@class:el-upload--text',timeout=3).click.to_upload('\n'.join(img_list))
            elif img_dir.name=="不动产抵押登记询问笔录":
                ele_annex_box.ele('tag=li@@text():询问笔录',timeout=3).ele('tag=div@@class:el-upload--text',timeout=3).click.to_upload('\n'.join(img_list))
            elif img_dir.name=="申请人身份证明":
                ele_annex_box.ele('tag=li@@text():身份证明',timeout=3).ele('tag=div@@class:el-upload--text',timeout=3).click.to_upload('\n'.join(img_list))
            elif img_dir.name=="委托函":
                ele_annex_box.ele('tag=li@@text():委托',timeout=3).ele('tag=div@@class:el-upload--text',timeout=3).click.to_upload('\n'.join(img_list))
            elif img_dir.name=="商品房预售合同":
                ele_annex_box.ele('tag=li@@text():商品房预售合同',timeout=3).ele('tag=div@@class:el-upload--text',timeout=3).click.to_upload('\n'.join(img_list))
            elif img_dir.name=="不动产登记证明":
                ele_annex_box.ele('tag=li@@text():不动产登记证明',timeout=3).ele('tag=div@@class:el-upload--text',timeout=3).click.to_upload('\n'.join(img_list))
            elif img_dir.name=="房产证":
                ele_annex_box.ele('tag=li@@text():不动产权证书或房屋所有权证',timeout=3).ele('tag=div@@class:el-upload--text',timeout=3).click.to_upload('\n'.join(img_list))
            elif img_dir.name=="借款合同":
                ele_annex_box.ele('tag=li@|text():借款合同@|text():授信合同@|text():债权合同',timeout=3).ele('tag=div@@class:el-upload--text',timeout=3).click.to_upload('\n'.join(img_list))
            elif img_dir.name=="抵押合同":
                ele_annex_box.ele('tag=li@@text():抵押合同',timeout=3).ele('tag=div@@class:el-upload--text',timeout=3).click.to_upload('\n'.join(img_list))
            elif img_dir.name=="不动产抵押权注销证明":
                ele_annex_box.ele('tag=li@@text():不动产抵押权注销证明',timeout=3).ele('tag=div@@class:el-upload--text',timeout=3).click.to_upload('\n'.join(img_list))
            elif img_dir.name=="关于抵押预告登记的约定":
                ele_annex_box.ele('tag=li@@text():关于预告抵押登记的约定',timeout=3).ele('tag=div@@class:el-upload--text',timeout=3).click.to_upload('\n'.join(img_list))
            elif img_dir.name=="结婚证或具结保证书":
                ele_annex_box.ele('tag=li@@text():结婚证或具结保证书',timeout=3).ele('tag=div@@class:el-upload--text',timeout=3).click.to_upload('\n'.join(img_list))
            elif img_dir.name=="其他":
                ele_annex_box.ele('tag=li@@text():其他',timeout=3).ele('tag=div@@class:el-upload--text',timeout=3).click.to_upload('\n'.join(img_list))
            ele_annex_box.wait(2)

#网页操作 - 保存
def webSave():
    tab=BROWSER.get_tab(url=URL)
    tab.ele('tag=span@@text():保存',timeout=3).click()

# async def main(businessinfo:BusinessInfo):
#    await asyncio.gather(webInputInfo(businessinfo),webUploadImg(businessinfo))



if __name__ == '__main__':
    # tab=Chromium().latest_tab
    # #添加抵押信息
    # ele_loan_info=tab.ele('tag=div@@class:yw_con@@text():担保范围')
    # ele_loan_term_start= ele_loan_info.ele('tag=label@@text():银行债务履行期限').parent().ele('tag=input@placeholder:开始日期',timeout=3)

    # ele_loan_term_start.click()



    business_type="合并登记"
    customers=[{"name":"张小华","id":"362422198609242011","phone":"12345678901"},{"name":"张小华","id":"362422198609242011","phone":"12345678901"},{"name":"张小华","id":"362422198609242011","phone":"12345678901"}]
    businessinfo=BusinessInfo(business_type=business_type,customers=customers)
    businessinfo.certificate=["赣（2024）吉安市不动产证明第7019815号"]
    businessinfo.loan_amount="1000000"
    businessinfo.loan_tream={"start":"20240101","end":"20241201"}
    businessinfo.loan_contract="2024000001"
    businessinfo.loan_guarantee="全部"
    businessinfo.img_path=r"C:\Users\Lcy\Desktop\抵质押登记\work\tempImg"
    
    #asyncio.run(writeBusinessInfoToWps(businessinfo,FILE_PATH))
    
    
    
    pass